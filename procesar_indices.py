#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
import glob
import re

# Intentar importar openpyxl, si no está instalado se auto-instala
try:
    import openpyxl
except ImportError:
    import subprocess
    import sys
    print("Instalando dependencia openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def limpiar_texto(val):
    if val is None:
        return ""
    # Convertir a string, quitar espacios en los extremos y normalizar espacios múltiples
    s = str(val).strip()
    s = re.sub(r'\s+', ' ', s)
    return s

def es_cabecera_o_titulo(val):
    if not val:
        return True
    s = val.upper().strip()
    
    # Si es una sola letra (cabecera de orden alfabético como 'A', 'B', etc.)
    if len(s) == 1 and s.isalpha():
        return True
        
    keywords = [
        "CONTRATANTES", "MATERIA", "REPERTORIO", "FOJAS", "RP", 
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    for kw in keywords:
        if kw in s:
            return True
    return False

def es_titulo_mes(val):
    if not val:
        return False
    s = val.upper().strip()
    meses = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    for m in meses:
        if m in s:
            return True
    return False

def procesar_excels():
    data_dir = "indices-data"
    output_dir = "indices-json"
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Creada carpeta de salida: {output_dir}")
        
    excel_files = glob.glob(os.path.join(data_dir, "AÑO *.xlsx"))
    
    if not excel_files:
        print("No se encontraron archivos Excel con el formato 'AÑO XXXX.xlsx' en 'indices-data/'.")
        return
        
    for filepath in excel_files:
        filename = os.path.basename(filepath)
        print(f"\nProcesando: {filename}")
        
        # Extraer el año del nombre del archivo (ej: AÑO 2026.xlsx -> 2026)
        match_year = re.search(r'\d{4}', filename)
        if not match_year:
            print(f"No se pudo determinar el año en el nombre del archivo: {filename}")
            continue
        year = match_year.group(0)
        
        # Cargar libro de Excel
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]] # Usar la primera hoja (consolidado del año)
        print(f"Leyendo hoja consolidada: {sheet.title}")
        
        registros = []
        mes_actual = ""
        
        for row in sheet.iter_rows(values_only=True):
            # En base al análisis, las columnas empiezan en la columna B (índice 1)
            # Fila típica: (None, Contratante 1, Contratante 2, Materia, Repertorio, Fojas, ...)
            if len(row) < 6:
                continue
                
            c1 = limpiar_texto(row[1])
            c2 = limpiar_texto(row[2])
            materia = limpiar_texto(row[3])
            repertorio = row[4]
            fojas = row[5]
            
            # Si es una fila que define un mes, guardamos el mes
            if c1 and es_titulo_mes(c1) and not c2 and not materia and not repertorio and not fojas:
                mes_actual = c1
                continue
            
            # Si el compareciente principal está vacío, omitir
            if not c1:
                continue
                
            # Si es una fila de cabecera o de título de mes, omitir
            if es_cabecera_o_titulo(c1):
                continue
                
            # Limpiar y validar números de repertorio y fojas
            rep_str = limpiar_texto(repertorio)
            foj_str = limpiar_texto(fojas)
            
            # Si no hay materia y no hay repertorio, probablemente no sea una escritura válida
            if not materia and not rep_str:
                continue
                
            # Crear registro estructurado
            reg = {
                "c1": c1,
                "c2": c2,
                "materia": materia,
                "repertorio": rep_str,
                "fojas": foj_str,
                "mes": mes_actual
            }
            registros.append(reg)
            
        wb.close()
        
        # Guardar en archivo JSON
        output_filepath = os.path.join(output_dir, f"indices-{year}.json")
        with open(output_filepath, "w", encoding="utf-8") as json_file:
            json.dump(registros, json_file, ensure_ascii=False, indent=2)
            
        print(f"¡Éxito! Creado '{output_filepath}' con {len(registros)} registros.")

if __name__ == "__main__":
    print("--- INICIANDO CONVERSIÓN DE ÍNDICES ---")
    procesar_excels()
    print("\n--- PROCESO TERMINADO ---")
